#!/usr/bin/env python3
"""
TikTok Search Scraper CLI
Author: TikTok Scraper Tool
Version: 1.0.0
"""

import asyncio
import aiohttp
import json
import csv
import re
import time
import uuid
import os
from datetime import datetime
from urllib.parse import urlencode
from typing import List, Dict, Optional
from dataclasses import dataclass, asdict

# Colors for terminal output
class Colors:
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'
    GRAY = '\033[90m'

def print_banner():
    """Print cool banner"""
    banner = f"""
{Colors.CYAN}╔══════════════════════════════════════════════════════════════╗
║                                                                  ║
║  {Colors.BOLD}{Colors.RED}T I K T O K {Colors.END}{Colors.CYAN}  {Colors.BOLD}S E A R C H   S C R A P E R{Colors.END}{Colors.CYAN}                         ║
║  {Colors.GRAY}━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━{Colors.CYAN}  ║
║  {Colors.GREEN}⚡ API-based scraper with signature generation{Colors.CYAN}                 ║
║  {Colors.GREEN}⚡ Export to JSON, CSV, or Excel format{Colors.CYAN}                       ║
║                                                                  ║
╚══════════════════════════════════════════════════════════════════╝{Colors.END}
    """
    print(banner)

def print_success(message: str):
    print(f"{Colors.GREEN}✅ {message}{Colors.END}")

def print_error(message: str):
    print(f"{Colors.RED}❌ {message}{Colors.END}")

def print_info(message: str):
    print(f"{Colors.BLUE}ℹ️  {message}{Colors.END}")

def print_warning(message: str):
    print(f"{Colors.YELLOW}⚠️  {message}{Colors.END}")

def print_progress(message: str):
    print(f"{Colors.CYAN}📡 {message}{Colors.END}")

@dataclass
class VideoData:
    """Data class for video information"""
    video_id: str
    url: str
    username: str
    nickname: str
    caption: str
    create_time: int
    duration: int
    plays: int
    likes: int
    comments: int
    shares: int
    collects: int
    followers: int
    music: str
    cover: str
    play_url: str

class TikTokAPIScraper:
    def __init__(self, signature_server_url: str = "http://localhost:8080"):
        self.signature_server = signature_server_url

    @staticmethod
    def _looks_relevant(items: List[Dict], keyword: str) -> bool:
        """Sesi tamu kadang diberi 'hasil pengganti' (feed fallback) yang tak
        ada kaitannya dengan kata kunci. Deteksi: adakah token kata kunci yang
        muncul di caption item. Hasil pengganti tidak pernah memuatnya."""
        tokens = [w.lower() for w in re.split(r"\W+", keyword)
                  if len(w) >= 3 and w.lower() not in {"yang", "dan", "dengan", "untuk", "dari", "ada"}]
        if not tokens:
            return True
        text = " ".join((i.get("desc") or "") for i in items[:8]).lower()
        return any(t in text for t in tokens)

    async def search_videos(self, keyword: str, max_videos: int = 30,
                           progress_callback=None) -> List[VideoData]:
        """Search TikTok videos menggunakan API"""

        videos = []
        cursor = 0
        search_id = f"{int(time.time() * 1000)}{uuid.uuid4().hex[:12].upper()}"

        async with aiohttp.ClientSession() as session:

            while len(videos) < max_videos:
                params = {
                    "aid": "1988",
                    "keyword": keyword,
                    "count": str(min(12, max_videos - len(videos))),
                    "cursor": str(cursor),
                    "search_source": "query",  # search_history memberi hasil tak relevan
                    "search_id": search_id,
                    "type": "1",  # wajib: filter video; tanpa ini TikTok beri hasil pengganti
                    "channel": "tiktok_web",
                }

                url = f"https://www.tiktok.com/api/search/general/full/?{urlencode(params)}"

                # TikTok menggilir hasil asli/hasil pengganti per-request di
                # sesi tamu; halaman tak relevan diulang sampai dapat yang asli
                # (pola alternasinya terbukti berbalik saat di-retry).
                retries = 0
                while True:
                    if progress_callback:
                        progress_callback(
                            f"Fetching page {cursor//12 + 1}..." + (f" (retry {retries})" if retries else ""),
                            len(videos))

                    async with session.post(
                        f"{self.signature_server}/fetch",
                        json={"url": url},
                        timeout=aiohttp.ClientTimeout(total=60)
                    ) as response:

                        result = await response.json()

                        if result.get("status") != "ok":
                            raise Exception(f"Fetch failed: {result}")

                        data = result.get("data", {})
                        items = self._parse_items(data)

                    if items and not self._looks_relevant(items, keyword) and retries < 3:
                        retries += 1
                        continue
                    break

                if not items:
                    break

                for item in items:
                    video = self._convert_to_video(item)
                    if video:
                        videos.append(video)

                cursor = data.get("cursor", cursor + len(items))

                if not data.get("has_more", False):
                    break

                await asyncio.sleep(1)

        return videos[:max_videos]

    def _parse_items(self, data: Dict) -> List[Dict]:
        items = []
        if not isinstance(data, dict):
            return items

        data_array = data.get("data")
        if isinstance(data_array, list):
            for entry in data_array:
                if isinstance(entry, dict):
                    item = entry.get("item")
                    if item and isinstance(item, dict):
                        items.append(item)
        return items

    def _convert_to_video(self, item: Dict) -> Optional[VideoData]:
        if not isinstance(item, dict):
            return None

        video_id = item.get("id")
        if not video_id:
            return None

        author = item.get("author", {})
        username = author.get("uniqueId", "") if isinstance(author, dict) else ""
        nickname = author.get("nickname", "") if isinstance(author, dict) else ""

        stats = item.get("stats", {})
        if not isinstance(stats, dict):
            stats = {}

        video_info = item.get("video", {})
        if not isinstance(video_info, dict):
            video_info = {}

        music = item.get("music", {})
        music_title = music.get("title", "") if isinstance(music, dict) else ""

        return VideoData(
            video_id=str(video_id),
            url=f"https://www.tiktok.com/@{username}/video/{video_id}" if username else "",
            username=username,
            nickname=nickname,
            caption=item.get("desc", ""),
            create_time=item.get("createTime", 0),
            duration=video_info.get("duration", 0),
            plays=int(stats.get("playCount", 0)),
            likes=int(stats.get("diggCount", 0)),
            comments=int(stats.get("commentCount", 0)),
            shares=int(stats.get("shareCount", 0)),
            collects=int(stats.get("collectCount", 0)),
            followers=int(author.get("followerCount", 0)) if isinstance(author, dict) else 0,
            music=music_title,
            cover=video_info.get("cover", ""),
            play_url=video_info.get("playAddr", "")
        )

    async def check_server_health(self) -> bool:
        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(
                    f"{self.signature_server}/health",
                    timeout=aiohttp.ClientTimeout(total=5)
                ) as response:
                    return response.status == 200
        except:
            return False


class Exporter:
    """Handle export to various formats"""

    @staticmethod
    def to_json(videos: List[VideoData], filepath: str):
        data = {
            "total": len(videos),
            "exported_at": datetime.now().isoformat(),
            "videos": [asdict(v) for v in videos]
        }
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    @staticmethod
    def to_csv(videos: List[VideoData], filepath: str):
        if not videos:
            return

        fieldnames = list(asdict(videos[0]).keys())
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for video in videos:
                writer.writerow(asdict(video))

    @staticmethod
    def to_excel(videos: List[VideoData], filepath: str):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "TikTok Videos"

            # Headers
            headers = list(asdict(videos[0]).keys()) if videos else []
            header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Data
            for row, video in enumerate(videos, 2):
                for col, (key, value) in enumerate(asdict(video).items(), 1):
                    ws.cell(row=row, column=col, value=value)

            # Auto-width columns
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15

            wb.save(filepath)
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")


class CLI:
    def __init__(self):
        self.scraper = TikTokAPIScraper("http://localhost:8080")
        self.exporter = Exporter()

    def clear_screen(self):
        os.system('cls' if os.name == 'nt' else 'clear')

    def get_user_input(self) -> tuple:
        """Get search parameters from user"""
        print(f"\n{Colors.BOLD}{Colors.YELLOW}📋 SEARCH PARAMETERS{Colors.END}")
        print(f"{Colors.GRAY}{'─' * 50}{Colors.END}")

        keyword = input(f"{Colors.CYAN}🔍 Keyword: {Colors.END}").strip()
        while not keyword:
            print_warning("Keyword cannot be empty!")
            keyword = input(f"{Colors.CYAN}🔍 Keyword: {Colors.END}").strip()

        while True:
            try:
                max_videos = input(f"{Colors.CYAN}📊 Number of videos (default 30): {Colors.END}").strip()
                max_videos = int(max_videos) if max_videos else 30
                if max_videos <= 0:
                    print_warning("Number must be positive!")
                    continue
                break
            except ValueError:
                print_warning("Please enter a valid number!")

        return keyword, max_videos

    def get_export_options(self) -> tuple:
        """Get export format and path from user"""
        print(f"\n{Colors.BOLD}{Colors.YELLOW}💾 EXPORT OPTIONS{Colors.END}")
        print(f"{Colors.GRAY}{'─' * 50}{Colors.END}")

        formats = {
            '1': ('json', 'JSON', Colors.GREEN),
            '2': ('csv', 'CSV', Colors.BLUE),
            '3': ('xlsx', 'Excel', Colors.CYAN)
        }

        print(f"\n{Colors.BOLD}Select format:{Colors.END}")
        for key, (ext, name, color) in formats.items():
            print(f"  {color}[{key}] {name} (.{ext}){Colors.END}")

        while True:
            choice = input(f"\n{Colors.CYAN}Format (1/2/3) [default: 1]: {Colors.END}").strip()
            if not choice:
                choice = '1'
            if choice in formats:
                ext, name, _ = formats[choice]
                break
            print_warning("Invalid choice! Please select 1, 2, or 3")

        # Default filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"tiktok_export_{timestamp}.{ext}"

        print(f"\n{Colors.BOLD}Save location:{Colors.END}")
        print(f"  {Colors.GRAY}[1] Current directory{Colors.END}")
        print(f"  {Colors.GRAY}[2] Custom path{Colors.END}")

        while True:
            loc_choice = input(f"\n{Colors.CYAN}Location (1/2) [default: 1]: {Colors.END}").strip()
            if not loc_choice:
                loc_choice = '1'
            if loc_choice in ['1', '2']:
                break
            print_warning("Invalid choice! Please select 1 or 2")

        if loc_choice == '1':
            filepath = default_filename
        else:
            custom_path = input(f"{Colors.CYAN}Enter full path: {Colors.END}").strip()
            if custom_path:
                if os.path.isdir(custom_path):
                    filepath = os.path.join(custom_path, default_filename)
                else:
                    filepath = custom_path
                    if not filepath.endswith(f'.{ext}'):
                        filepath += f'.{ext}'
            else:
                filepath = default_filename

        return ext, filepath

    def display_results(self, videos: List[VideoData], keyword: str, elapsed: float):
        """Display search results in a nice format"""
        self.clear_screen()
        print_banner()

        print(f"\n{Colors.BOLD}{Colors.GREEN}✅ SEARCH COMPLETED{Colors.END}")
        print(f"{Colors.GRAY}{'─' * 50}{Colors.END}")
        print(f"  {Colors.CYAN}Keyword:{Colors.END} {keyword}")
        print(f"  {Colors.CYAN}Videos found:{Colors.END} {len(videos)}")
        print(f"  {Colors.CYAN}Time elapsed:{Colors.END} {elapsed:.2f} seconds")

        if videos:
            print(f"\n{Colors.BOLD}{Colors.YELLOW}📹 TOP RESULTS{Colors.END}")
            print(f"{Colors.GRAY}{'─' * 70}{Colors.END}")

            for i, v in enumerate(videos[:5], 1):
                print(f"\n{Colors.BOLD}{i}. @{v.username}{Colors.END} {Colors.GRAY}({v.nickname}){Colors.END}")
                if v.caption:
                    caption = v.caption[:60] + "..." if len(v.caption) > 60 else v.caption
                    print(f"   {Colors.GRAY}📝{Colors.END} {caption}")
                print(f"   {Colors.GREEN}▶️ {v.plays:,}{Colors.END} views  {Colors.RED}❤️ {v.likes:,}{Colors.END} likes")
                print(f"   {Colors.BLUE}🔗{Colors.END} {v.url}")

    async def run(self):
        """Main CLI loop"""
        self.clear_screen()
        print_banner()

        # Check server health
        print(f"\n{Colors.GRAY}Checking signature server...{Colors.END}")
        if not await self.scraper.check_server_health():
            print_error("Signature server is not running!")
            print_warning("\nPlease start the signature server first:")
            print(f"{Colors.GRAY}  cd tiktok-signature && npm start{Colors.END}\n")
            return

        print_success("Signature server is ready!")

        while True:
            # Get search parameters
            keyword, max_videos = self.get_user_input()

            # Start search
            print(f"\n{Colors.BOLD}{Colors.YELLOW}🚀 STARTING SEARCH{Colors.END}")
            print(f"{Colors.GRAY}{'─' * 50}{Colors.END}")

            start_time = time.time()

            def progress_callback(message: str, current_count: int):
                print(f"{Colors.CYAN}  {message} {Colors.GRAY}({current_count} videos collected){Colors.END}")

            try:
                videos = await self.scraper.search_videos(
                    keyword, max_videos, progress_callback
                )
                elapsed = time.time() - start_time

                if not videos:
                    print_error("No videos found!")
                    retry = input(f"\n{Colors.CYAN}Try another keyword? (y/n): {Colors.END}").strip().lower()
                    if retry == 'y':
                        self.clear_screen()
                        print_banner()
                        continue
                    else:
                        break

                # Display results
                self.display_results(videos, keyword, elapsed)

                # Export options
                export_choice = input(f"\n{Colors.CYAN}Export results? (y/n): {Colors.END}").strip().lower()

                if export_choice == 'y':
                    ext, filepath = self.get_export_options()

                    print(f"\n{Colors.GRAY}Exporting to {ext.upper()}...{Colors.END}")

                    try:
                        # Create directory if needed
                        os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)

                        if ext == 'json':
                            self.exporter.to_json(videos, filepath)
                        elif ext == 'csv':
                            self.exporter.to_csv(videos, filepath)
                        elif ext == 'xlsx':
                            self.exporter.to_excel(videos, filepath)

                        print_success(f"Successfully exported to: {filepath}")

                        # Show file size
                        size = os.path.getsize(filepath)
                        if size < 1024:
                            size_str = f"{size} B"
                        elif size < 1024 * 1024:
                            size_str = f"{size / 1024:.2f} KB"
                        else:
                            size_str = f"{size / (1024 * 1024):.2f} MB"
                        print(f"{Colors.GRAY}  File size: {size_str}{Colors.END}")

                    except ImportError as e:
                        print_error(str(e))
                    except Exception as e:
                        print_error(f"Export failed: {e}")

                # Ask for another search
                print(f"\n{Colors.GRAY}{'─' * 50}{Colors.END}")
                again = input(f"{Colors.CYAN}Perform another search? (y/n): {Colors.END}").strip().lower()
                if again == 'y':
                    self.clear_screen()
                    print_banner()
                    continue
                else:
                    print(f"\n{Colors.GREEN}👋 Thank you for using TikTok Scraper!{Colors.END}\n")
                    break

            except KeyboardInterrupt:
                print(f"\n\n{Colors.YELLOW}⚠️  Search interrupted by user{Colors.END}")
                break
            except Exception as e:
                print_error(f"Error: {e}")
                break


async def main():
    cli = CLI()
    await cli.run()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print(f"\n\n{Colors.YELLOW}👋 Goodbye!{Colors.END}\n")
    except Exception as e:
        print_error(f"Fatal error: {e}")
